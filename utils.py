import streamlit as st
import os
import base64
import openpyxl

from openpyxl import Workbook

import google.generativeai as genai

from dotenv import load_dotenv

load_dotenv()

genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))



def get_gemini_responses(input, image, prompts):
    model = genai.GenerativeModel('gemini-1.5-flash',
                                  generation_config=genai.types.GenerationConfig(
                                      temperature=0.5,
                                      top_p=0.9,
                                      top_k=40,
                                      max_output_tokens=1024
                                  ))
    all_responses = []
    for prompt in prompts:
        response = model.generate_content([input, image[0], prompt])
        for part in response.parts:
            if part.text:
                all_responses.append(part.text)
    #print(all_responses)
    return all_responses

def get_gemini_dims_responses(input, image, prompts):
    model = genai.GenerativeModel('gemini-1.5-flash',
                                  generation_config=genai.types.GenerationConfig(
                                      temperature=0.5,
                                      top_p=0.9,
                                      top_k=40,
                                      max_output_tokens=512
                                  ))
    all_responses = []
    for prompt in prompts:
        response = model.generate_content([input, image[0], prompt])
        # print(response)
        for part in response.parts:
            if part.text:
                all_responses.append(part.text)
    # print(all_responses)
    return all_responses[0]

# def input_image_setup(uploaded_file):
#     if uploaded_file is not None:
#         bytes_data = uploaded_file.getvalue()
#         image_parts = [{"mime_type": uploaded_file.type, "data": bytes_data}]
#         return image_parts
#     else:
#         raise FileNotFoundError("No file uploaded")
    
def input_image_setup(file_bytes_io, mime_type):
    if file_bytes_io is not None:
        bytes_data = file_bytes_io.getvalue()
        image_parts = [{"mime_type": mime_type, "data": bytes_data}]
        return image_parts
    else:
        raise FileNotFoundError("No file uploaded")
    
def input_image_setup_local(image_path):
    with open(image_path, "rb") as img_file:
        bytes_data = img_file.read()
        mime_type = "image/jpeg" if image_path.endswith(".jpg") or image_path.endswith(".jpeg") else "image/png"
        image_parts = [{"mime_type": mime_type, "data": bytes_data}]
        return image_parts

def encode_image(uploaded_image):
    img_bytes = uploaded_image.read()
    return base64.b64encode(img_bytes).decode("utf-8")



def write_to_excel_meesho(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[1]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[3] if cell.value}
        headers = {key.strip().split('\n\n')[0]: value for key, value in headers.items()}

        #print(headers)
        #print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    # print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=1).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1

    for image_name, response , description in results:
        #print(response)
        # answers = response.split("\n")  
        # answers = [ans.strip() for ans in answers]
        # answers.insert(0, image_name)

        # print()
        # print(answers)  
        # print()

        # Create a dictionary mapping field names to values
        # print(target_fields)
        # field_values = {field: answers[i] if i < len(answers) else "" for i, field in enumerate(target_fields)}
        # print()
        # print(field_values)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["Product Description"] = description
        field_values["Fields + Description:"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)


def write_to_excel_flipkart(results, filename, target_fields, fixed_values):
    abs_path = os.path.abspath(filename)
    print(f"[DEBUG] write_to_excel_flipkart: absolute path = {abs_path}")

    # 1) Load or create
    if os.path.exists(filename):
        print("[DEBUG] File exists → loading workbook.")
        try:
            wb = openpyxl.load_workbook(filename)
        except Exception as e:
            print(f"[ERROR] Could not load workbook: {e}")
            raise

        # 2) Show all sheet names
        print(f"[DEBUG] Workbook sheets: {wb.sheetnames!r}")

        # 3) Pick the correct Flipkart sheet
        if "earring" in wb.sheetnames:
            ws = wb["earring"]
            print("[DEBUG] Selected sheet by name: 'earring'.")
        else:
            # If you still have code that falls back to wb.active, print that
            ws = wb.active
            print(f"[DEBUG] 'earring' not found. Using active sheet: '{ws.title}'.")

        # 4) Read the header row (row 1)
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        print(f"[DEBUG] Read headers from row 1: {list(headers.keys())}\n")

    else:
        print("[DEBUG] File does NOT exist → creating brand-new workbook.")
        wb = Workbook()
        ws = wb.active
        ws.title = "earring"
        print("[DEBUG] Created new sheet and renamed to 'earring'.")

        default_headers = [
            "Flipkart Serial Number", "Seller SKU ID", "Listing Status",
            "MRP (INR)", "Your selling price (INR)", "Fullfilment by",
            "Procurement type", "Procurement SLA (DAY)", "Stock",
            "Shipping provider", "Local delivery charge (INR)",
            "Zonal delivery charge (INR)", "National delivery charge (INR)",
            "Height (CM)", "Weight (KG)", "Breadth (CM)", "Length (CM)",
            "HSN", "Luxury Cess", "Country Of Origin", "Manufacturer Details",
            "Packer Details", "Importer Details", "Tax Code",
            "Minimum Order Quantity (MinOQ)", "Brand", "Model Number",
            "Type", "Color", "Ideal For", "Model Name", "Base Material",
            "Gemstone", "Diamond Color Grade", "Diamond Clarity",
            "Pearl Type", "Certification", "Collection", "Plating",
            "Silver Weight (g)", "Diamond Shape", "Diamond Weight (carat)",
            "Main Image URL", "Other Image URL 1", "Other Image URL 2",
            "Other Image URL 3", "Other Image URL 4", "Group ID",
            "Occasion", "EAN/UPC", "EAN/UPC - Measuring Unit",
            "Piercing Required", "Number of Pairs", "Number of Gemstones",
            "Earring Back Type", "Finish", "Setting", "Design",
            "Silver Purity", "Silver Color", "Metal Purity", "Metal Color",
            "Metal Weight", "Natural/Synthetic Diamond",
            "Diamond Width (mm)", "Diamond Height (mm)", "Natural/Synthetic Ruby",
            "Ruby Shape", "Ruby Clarity", "Ruby Color", "Ruby Width (mm)",
            "Ruby Height (mm)", "Ruby Weight (carat)",
            "Natural/Synthetic Emerald", "Emerald Shape", "Emerald Clarity",
            "Emerald Color", "Emerald Width (mm)", "Emerald Height (mm)",
            "Emerald Weight (carat)", "Natural/Synthetic Sapphire",
            "Sapphire Shape", "Sapphire Clarity", "Sapphire Color",
            "Sapphire Width (mm)", "Sapphire Height (mm)",
            "Sapphire Weight (carat)", "Natural/Synthetic Amethyst",
            "Amethyst Shape", "Amethyst Clarity", "Amethyst Color",
            "Amethyst Width (mm)", "Amethyst Height (mm)", "Amethyst Weight (carat)",
            "Artificial Pearl Material", "Pearl Shape", "Pearl Grade",
            "Pearl Color", "Pearl Length (mm)", "Pearl Diameter (mm)",
            "Natural/Synthetic Semi-precious Stone", "Semi-precious Stone Type",
            "Semi-precious Stone Shape", "Width (mm)", "Height (mm)",
            "Diameter (mm)", "Weight (g)", "Other Dimensions", "Other Features",
            "Sales Package", "Description", "Search Keywords", "Key Features",
            "Video URL", "Domestic Warranty", "Domestic Warranty - Measuring Unit",
            "International Warranty", "International Warranty - Measuring Unit",
            "External Identifier", "Trend", "Warranty Summary",
            "Warranty Service Type", "Covered in Warranty", "Not Covered in Warranty",
            "Closure Type", "Sub Type", "Earring Shape", "With Ear Chain",
            "Earring Set Type", "Ornamentation Type", "Trend AW 16",
            "Net Quantity", "Brand Color", "Supplier Image"
        ]
        for idx, h in enumerate(default_headers, start=1):
            ws.cell(row=1, column=idx, value=h)
        headers = {name: i for i, name in enumerate(default_headers, start=1)}
        print(f"[DEBUG] Wrote default headers in new workbook: {list(headers.keys())}\n")

    # 5) Print out exactly what target_fields and fixed_values contain
    print(f"[DEBUG] target_fields passed in: {target_fields}\n")
    print(f"[DEBUG] fixed_values passed in: {list(fixed_values.keys())}\n")

    # 6) Build “all_fields” and see which actually exist in headers
    all_fields = set(target_fields) | set(fixed_values.keys())
    actual_columns = {f: headers[f] for f in all_fields if f in headers}
    missing_columns = [f for f in all_fields if f not in headers]

    print(f"[DEBUG] target_columns (field → column#) = {actual_columns}\n")
    print(f"[DEBUG] missing_fields = {missing_columns}\n")

    # 7) Find first empty row by checking “Seller SKU ID” (column 7):
    if "Seller SKU ID" not in headers:
        raise RuntimeError("Column 'Seller SKU ID' not found in header.")
    sku_col = headers["Seller SKU ID"]
    row_idx = 2
    while True:
        val = ws.cell(row=row_idx, column=sku_col).value
        if val is None or str(val).strip() == "":
            print(f"[DEBUG] First empty row for writing = {row_idx}\n")
            break
        row_idx += 1

    # 8) Write each result
    for image_name, response, description in results:
        field_values = dict(response)
        for k, v in fixed_values.items():
            field_values[k] = v
        field_values["Description"] = description
        field_values["Seller SKU ID"] = os.path.splitext(image_name)[0]

        for field, col_idx in actual_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        print(f"[DEBUG] Wrote row {row_idx}: "
              f"{ {f: field_values.get(f) for f in actual_columns} }")
        row_idx += 1

    # 9) Save and report
    try:
        wb.save(filename)
        print(f"[DEBUG] Saved workbook successfully to '{abs_path}'\n")
    except PermissionError:
        print(f"[ERROR] Permission denied when saving '{abs_path}'")
        raise
    except Exception as e:
        print(f"[ERROR] Unexpected save error: {e}")
        raise

# def write_to_excel_flipkart(results,filename,target_fields,fixed_values):

#     if os.path.exists(filename):
#         print("started loading")
#         print()
#         wb = openpyxl.load_workbook(filename)
#         #print("finished")
#         #print()
#         ws = wb.worksheets[2]

#         # Read headers dynamically from row 1
#         headers = {cell.value: cell.column for cell in ws[1] if cell.value}
#         print(headers)
#         #print()
#     else:
#         print("not found")
#         wb = Workbook()
#         ws = wb.active  

#         # Define default headers
#         default_headers = [
#             "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
#             "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
#             "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
#             "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
#             "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
#             "Ornamentation Type", "Trend"
#         ]
        
#         headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
#         ws.append(default_headers)  # Write headers in row 1 if creating a new file

#     #print()
#     # Combine target fields and fixed value fields
#     all_fields = set(target_fields).union(fixed_values.keys())

#     # Get column indices for all relevant fields
#     target_columns = {field: headers[field] for field in all_fields if field in headers}
#     # print(target_columns)

#     # **Write output from row 6 onwards**
#     row_idx = 1
#     while ws.cell(row=row_idx, column=7).value:  # Assuming column 1 is "Image Name" or always filled
#         row_idx += 1
#     #print(row_idx)  
#     for image_name, response , description in results:
#         print(response)
#         field_values = response

#         # Write values only in the specified fields
#         for key, val in fixed_values.items():
#             field_values[key] = val

#         field_values["Description"] = description
#         field_values["Seller SKU ID"] = os.path.splitext(image_name)[0]

#         for field, col_idx in target_columns.items():
#             ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

#         row_idx += 1  # Move to the next row

#     wb.save(filename)


def write_to_excel_amz_xl(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[0]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[3] if cell.value}
        #print(headers)
        #print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    #print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=2).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1
    #print(row_idx)  
    for image_name, response , description in results:
        # print(response)
        # answers = response.split("\n")  
        # answers = [ans.strip() for ans in answers]
        # answers.insert(0, os.path.splitext(image_name)[0])

        # print()
        # print(answers)  
        # print()

        # Create a dictionary mapping field names to values
        # print(target_fields)
        # field_values = {field: answers[i] if i < len(answers) else "" for i, field in enumerate(target_fields)}
        # print()
        # print(field_values)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["product_description"] = description
        field_values["item_sku"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)